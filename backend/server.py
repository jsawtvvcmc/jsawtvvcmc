from fastapi import FastAPI, APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
import jwt
from jwt.exceptions import ExpiredSignatureError, InvalidTokenError
from enum import Enum
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import httpx

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI(title="ABC Program Management System")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Security
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()
SECRET_KEY = os.environ.get("JWT_SECRET_KEY", "your-secret-key-change-in-production")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7  # 7 days

# Enums
class UserRole(str, Enum):
    SUPER_ADMIN = "Super Admin"  # Global access to all projects
    ADMIN = "Admin"              # Project-level admin
    DRIVER = "Driver"
    CATCHER = "Catcher"
    VETERINARY = "Veterinary Doctor"
    CARETAKER = "Caretaker"

class ProjectStatus(str, Enum):
    ACTIVE = "Active"
    INACTIVE = "Inactive"
    SUSPENDED = "Suspended"

class CaseStatus(str, Enum):
    CAUGHT = "Caught"
    IN_KENNEL = "In Kennel"
    SURGERY_SCHEDULED = "Surgery Scheduled"
    SURGERY_COMPLETED = "Surgery Completed"
    SURGERY_CANCELLED = "Surgery Cancelled"
    UNDER_TREATMENT = "Under Treatment"
    READY_FOR_RELEASE = "Ready for Release"
    DISPATCHED = "Dispatched"
    RELEASED = "Released"
    DECEASED = "Deceased"

class Gender(str, Enum):
    MALE = "Male"
    FEMALE = "Female"

class CancellationReason(str, Enum):
    TOO_WEAK = "Too weak"
    UNDER_AGE = "Under age"
    LOOKS_ILL = "Looks ill"
    CONTAGIOUS_DISEASE = "Shows symptoms of highly contagious disease"
    ALREADY_STERILIZED = "Already sterilized"
    ADVANCED_PREGNANT = "Advanced pregnant"
    LACTATING = "Lactating"
    OTHER = "Other"

# Medicine Protocol - Standard dosages based on weight (per 10kg baseline)
MEDICINE_PROTOCOL = {
    "Anti-Rabies Vaccine": {"base_dose": 1, "unit": "Ml", "fixed": True},
    "Xylazine": {"base_dose": 1, "unit": "Ml", "per_10kg": True},
    "Melonex": {"base_dose": 0.8, "unit": "Ml", "per_10kg": True, "max": 1},
    "Atropine": {"base_dose": 1, "unit": "Ml", "per_10kg": True, "round_half": True},
    "Diazepam": {"base_dose": 0, "unit": "Ml", "per_10kg": True},
    "Prednisolone": {"base_dose": 0, "unit": "Ml", "per_10kg": True},
    "Ketamine": {"base_dose": 3, "unit": "Ml", "per_10kg": True, "round_half": True},
    "Tribivet": {"base_dose": 1, "unit": "Ml", "fixed": True},
    "Intacef Tazo": {"base_dose": 400, "unit": "Mg", "per_10kg": True, "round_50": True},
    "Adrenaline": {"base_dose": 0, "unit": "Ml", "per_10kg": True},
    "Alu Spray": {"base_dose": 2, "unit": "Ml", "per_10kg": True, "round_half": True},
    "Ethamsylate": {"base_dose": 1, "unit": "Ml", "per_10kg": True, "round_half": True},
    "Tincture": {"base_dose": 20, "unit": "Ml", "per_10kg": True, "round_5": True},
    "Avil": {"base_dose": 1, "unit": "Ml", "fixed": True},
    "Vicryl 1": {"base_dose": 0.20, "unit": "Pcs", "fixed": True},
    "Catgut": {"base_dose": 0.20, "unit": "Pcs", "fixed": True},
    "Vicryl 2": {"base_dose": 0.20, "unit": "Pcs", "female_only": True},
    "Metronidazole": {"base_dose": 50, "unit": "Ml", "fixed": True},
}

def calculate_medicine_dosage(weight: float, medicine_name: str, gender: str = None) -> float:
    """Calculate medicine dosage based on weight and protocol"""
    if medicine_name not in MEDICINE_PROTOCOL:
        return 0
    
    protocol = MEDICINE_PROTOCOL[medicine_name]
    
    # Check female-only medicines
    if protocol.get("female_only") and gender != "Female":
        return 0
    
    base = protocol["base_dose"]
    
    if protocol.get("fixed"):
        return base
    
    if protocol.get("per_10kg"):
        dose = base * weight / 10
        
        # Apply max limit
        if protocol.get("max"):
            dose = min(dose, protocol["max"])
        
        # Apply rounding
        if protocol.get("round_half"):
            dose = round(dose * 2) / 2
        elif protocol.get("round_50"):
            dose = round(dose / 50) * 50
        elif protocol.get("round_5"):
            dose = round(dose / 5) * 5
        
        return dose
    
    return base

# Helper functions
def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def decode_token(token: str) -> dict:
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return payload
    except ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired")
    except InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    token = credentials.credentials
    payload = decode_token(token)
    user_id = payload.get("user_id")
    if not user_id:
        raise HTTPException(status_code=401, detail="Invalid authentication credentials")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user

def require_roles(allowed_roles: List[UserRole]):
    async def role_checker(current_user: dict = Depends(get_current_user)) -> dict:
        if current_user["role"] not in [role.value for role in allowed_roles]:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return current_user
    return role_checker

# Models
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    first_name: str
    last_name: str
    mobile: str
    role: UserRole
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    email: EmailStr
    first_name: str
    last_name: str
    mobile: str
    role: UserRole

class UserInDB(User):
    password_hash: str

class LoginRequest(BaseModel):
    email: EmailStr
    password: str

class LoginResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: User

# Startup event to create default super admin and configuration
@app.on_event("startup")
async def create_default_superuser():
    """Create default super admin if not exists"""
    default_email = "manoj@janicestrust.org"
    existing_user = await db.users.find_one({"email": default_email}, {"_id": 0})
    
    if not existing_user:
        default_user = {
            "id": str(uuid.uuid4()),
            "email": default_email,
            "first_name": "Manoj",
            "last_name": "Oswal",
            "mobile": "9890044455",
            "role": UserRole.SUPER_ADMIN.value,  # Global Super Admin
            "project_id": None,  # None = access to all projects
            "password_hash": hash_password("Kashid@25067"),
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(default_user)
        logger.info(f"Default super admin created: {default_email}")
    else:
        # Update existing user to Super Admin if needed
        if existing_user.get("role") == "Super User":
            await db.users.update_one(
                {"email": default_email},
                {"$set": {"role": UserRole.SUPER_ADMIN.value, "project_id": None}}
            )
            logger.info(f"Updated user to Super Admin: {default_email}")
        logger.info(f"Default super admin already exists: {default_email}")
    
    # Create default system configuration (legacy - for backward compatibility)
    existing_config = await db.system_config.find_one({"id": "system_config"}, {"_id": 0})
    if not existing_config:
        default_config = {
            "id": "system_config",
            "organization_name": "Janice Smith Animal Welfare Trust",
            "organization_shortcode": "JS",
            "registered_office": "352, Vadgaon, Yashwant Nagar, Talegaon Dabhade, Maharashtra 410507",
            "organization_logo": None,
            "project_name": "Talegaon ABC Project",
            "project_code": "TAL",
            "municipal_logo": None,
            "project_address": "352, Vadgaon, Yashwant Nagar, Talegaon Dabhade, Maharashtra 410507, India",
            "max_kennels": 300,
            "cloud_provider": "google_drive",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.system_config.insert_one(default_config)
        logger.info("Default system configuration created")
    
    # Create default project if none exists
    existing_projects = await db.projects.count_documents({})
    if existing_projects == 0:
        default_project = {
            "id": str(uuid.uuid4()),
            "organization_name": "Janice Smith Animal Welfare Trust",
            "organization_shortcode": "JS",
            "organization_logo_url": None,
            "project_name": "Talegaon ABC Project",
            "project_code": "TAL",
            "project_logo_url": None,
            "project_address": "352, Vadgaon, Yashwant Nagar, Talegaon Dabhade, Maharashtra 410507, India",
            "max_kennels": 300,
            "status": ProjectStatus.ACTIVE.value,
            "drive_folder_id": None,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.projects.insert_one(default_project)
        logger.info("Default project (TAL) created")
    
    # Create default medicines for surgery (global - will be copied to projects)
    existing_medicines_count = await db.medicines.count_documents({})
    if existing_medicines_count == 0:
        default_medicines = [
            {"name": "Anti-Rabies Vaccine", "unit": "Ml", "packing": "Vial", "packing_size": 10},
            {"name": "Xylazine", "unit": "Ml", "packing": "Vial", "packing_size": 30},
            {"name": "Melonex", "unit": "Ml", "packing": "Bottle", "packing_size": 100},
            {"name": "Atropine", "unit": "Ml", "packing": "Vial", "packing_size": 30},
            {"name": "Diazepam", "unit": "Ml", "packing": "Vial", "packing_size": 10},
            {"name": "Prednisolone", "unit": "Ml", "packing": "Vial", "packing_size": 10},
            {"name": "Ketamine", "unit": "Ml", "packing": "Vial", "packing_size": 5},
            {"name": "Tribivet", "unit": "Ml", "packing": "Bottle", "packing_size": 100},
            {"name": "Intacef Tazo", "unit": "Mg", "packing": "Vial", "packing_size": 4500},
            {"name": "Adrenaline", "unit": "Ml", "packing": "Vial", "packing_size": 10},
            {"name": "Alu Spray", "unit": "Ml", "packing": "Bottle", "packing_size": 100},
            {"name": "Ethamsylate", "unit": "Ml", "packing": "Vial", "packing_size": 30},
            {"name": "Tincture", "unit": "Ml", "packing": "Bottle", "packing_size": 400},
            {"name": "Avil", "unit": "Ml", "packing": "Bottle", "packing_size": 100},
            {"name": "Vicryl 1", "unit": "Pcs", "packing": "Pack", "packing_size": 12},
            {"name": "Catgut", "unit": "Pcs", "packing": "Pack", "packing_size": 12},
            {"name": "Vicryl 2", "unit": "Pcs", "packing": "Pack", "packing_size": 12},
            {"name": "Metronidazole", "unit": "Ml", "packing": "Bottle", "packing_size": 100},
        ]
        
        for med in default_medicines:
            medicine_doc = {
                "id": str(uuid.uuid4()),
                "name": med["name"],
                "generic_name": None,
                "unit": med["unit"],
                "packing": med["packing"],
                "packing_size": med["packing_size"],
                "current_stock": 0.0,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.medicines.insert_one(medicine_doc)
        
        logger.info(f"Created {len(default_medicines)} default medicines for surgery")

# Authentication Routes
@api_router.post("/auth/login", response_model=LoginResponse)
async def login(login_data: LoginRequest):
    """Login endpoint"""
    user = await db.users.find_one({"email": login_data.email}, {"_id": 0})
    
    if not user or not verify_password(login_data.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if not user["is_active"]:
        raise HTTPException(status_code=403, detail="User account is inactive")
    
    # Create access token
    access_token = create_access_token({"user_id": user["id"], "email": user["email"]})
    
    # Remove password hash before returning
    user.pop("password_hash", None)
    if isinstance(user['created_at'], str):
        user['created_at'] = datetime.fromisoformat(user['created_at'])
    
    return LoginResponse(access_token=access_token, user=User(**user))

@api_router.get("/auth/me", response_model=User)
async def get_current_user_info(current_user: dict = Depends(get_current_user)):
    """Get current user information"""
    if isinstance(current_user['created_at'], str):
        current_user['created_at'] = datetime.fromisoformat(current_user['created_at'])
    return User(**current_user)

@api_router.get("/config")
async def get_system_config():
    """Get system configuration"""
    config = await db.system_config.find_one({"id": "system_config"}, {"_id": 0})
    if not config:
        return {
            "id": "system_config",
            "organization_name": "Janice Smith Animal Welfare Trust",
            "organization_shortcode": "JS",
            "registered_office": "352, Vadgaon, Yashwant Nagar, Talegaon Dabhade, Maharashtra 410507",
            "organization_logo": None,
            "project_name": "Talegaon ABC Project",
            "project_code": "TAL",
            "municipal_logo": None,
            "project_address": "352, Vadgaon, Yashwant Nagar, Talegaon Dabhade, Maharashtra 410507, India",
            "max_kennels": 300,
            "cloud_provider": "google_drive"
        }
    return config

@api_router.put("/config")
async def update_system_config(
    config_data: dict,
    current_user: dict = Depends(require_roles([UserRole.SUPER_USER, UserRole.ADMIN]))
):
    """Update system configuration"""
    # Fields that can be updated
    allowed_fields = [
        "organization_name", "organization_shortcode", "registered_office",
        "organization_logo", "project_name", "project_code", "municipal_logo",
        "project_address", "max_kennels", "cloud_provider", "google_maps_api_key"
    ]
    
    update_data = {k: v for k, v in config_data.items() if k in allowed_fields}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # Validate max_kennels
    if "max_kennels" in update_data:
        max_kennels = update_data["max_kennels"]
        if not isinstance(max_kennels, int) or max_kennels < 1 or max_kennels > 300:
            raise HTTPException(status_code=400, detail="max_kennels must be between 1 and 300")
        
        # Update kennels collection if max_kennels changed
        current_config = await db.system_config.find_one({"id": "system_config"}, {"_id": 0})
        current_max = current_config.get("max_kennels", 300) if current_config else 300
        
        if max_kennels != current_max:
            # Remove extra kennels if reducing
            if max_kennels < current_max:
                await db.kennels.delete_many({"kennel_number": {"$gt": max_kennels}})
            # Add new kennels if increasing
            elif max_kennels > current_max:
                for i in range(current_max + 1, max_kennels + 1):
                    existing = await db.kennels.find_one({"kennel_number": i})
                    if not existing:
                        await db.kennels.insert_one({
                            "id": str(uuid.uuid4()),
                            "kennel_number": i,
                            "is_occupied": False,
                            "current_case_id": None,
                            "last_updated": datetime.now(timezone.utc).isoformat()
                        })
    
    # Validate shortcodes (uppercase, 2-5 chars)
    if "organization_shortcode" in update_data:
        shortcode = update_data["organization_shortcode"].upper()
        if len(shortcode) < 2 or len(shortcode) > 5:
            raise HTTPException(status_code=400, detail="Organization shortcode must be 2-5 characters")
        update_data["organization_shortcode"] = shortcode
    
    if "project_code" in update_data:
        project_code = update_data["project_code"].upper()
        if len(project_code) < 2 or len(project_code) > 5:
            raise HTTPException(status_code=400, detail="Project code must be 2-5 characters")
        update_data["project_code"] = project_code
    
    await db.system_config.update_one(
        {"id": "system_config"},
        {"$set": update_data},
        upsert=True
    )
    
    return {"message": "Configuration updated successfully"}

@api_router.post("/config/logo/{logo_type}")
async def upload_logo(
    logo_type: str,
    data: dict,
    current_user: dict = Depends(require_roles([UserRole.SUPER_USER, UserRole.ADMIN]))
):
    """Upload organization or municipal logo (base64)"""
    if logo_type not in ["organization", "municipal"]:
        raise HTTPException(status_code=400, detail="logo_type must be 'organization' or 'municipal'")
    
    logo_data = data.get("logo_base64")
    if not logo_data:
        raise HTTPException(status_code=400, detail="logo_base64 is required")
    
    field_name = f"{logo_type}_logo"
    
    await db.system_config.update_one(
        {"id": "system_config"},
        {"$set": {field_name: logo_data, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    
    return {"message": f"{logo_type.capitalize()} logo uploaded successfully"}

# Basic route
@api_router.get("/")
async def root():
    return {"message": "ABC Program Management System API", "version": "1.0"}

# ==================== GEOCODING ====================

@api_router.get("/geocode/reverse")
async def reverse_geocode(
    lat: float,
    lng: float,
    current_user: dict = Depends(get_current_user)
):
    """
    Convert latitude/longitude to address.
    Uses Google Maps API if configured, otherwise falls back to OpenStreetMap.
    """
    try:
        # Check if Google Maps API key is configured
        config = await db.system_config.find_one({"id": "system_config"}, {"_id": 0})
        google_maps_key = config.get("google_maps_api_key") if config else None
        
        if google_maps_key:
            # Use Google Maps Geocoding API
            url = "https://maps.googleapis.com/maps/api/geocode/json"
            params = {
                "latlng": f"{lat},{lng}",
                "key": google_maps_key,
                "language": "en"
            }
            
            async with httpx.AsyncClient() as client:
                response = await client.get(url, params=params, timeout=10.0)
                
                if response.status_code == 200:
                    data = response.json()
                    
                    if data.get("status") == "OK" and data.get("results"):
                        result = data["results"][0]
                        address = result.get("formatted_address", "")
                        
                        # Extract address components
                        components = {}
                        for component in result.get("address_components", []):
                            types = component.get("types", [])
                            if "route" in types:
                                components["road"] = component["long_name"]
                            elif "sublocality" in types or "neighborhood" in types:
                                components["suburb"] = component["long_name"]
                            elif "locality" in types:
                                components["city"] = component["long_name"]
                            elif "administrative_area_level_1" in types:
                                components["state"] = component["long_name"]
                            elif "postal_code" in types:
                                components["postcode"] = component["long_name"]
                            elif "country" in types:
                                components["country"] = component["long_name"]
                        
                        return {
                            "success": True,
                            "address": address,
                            "formatted_address": address,
                            "components": components,
                            "lat": lat,
                            "lng": lng,
                            "provider": "google_maps"
                        }
                    elif data.get("status") == "REQUEST_DENIED":
                        logger.error(f"Google Maps API error: {data.get('error_message', 'Request denied')}")
                    elif data.get("status") == "OVER_QUERY_LIMIT":
                        logger.warning("Google Maps API quota exceeded, falling back to OpenStreetMap")
        
        # Fallback to OpenStreetMap/Nominatim (free)
        url = "https://nominatim.openstreetmap.org/reverse"
        params = {
            "lat": lat,
            "lon": lng,
            "format": "json",
            "addressdetails": 1
        }
        headers = {
            "User-Agent": "J-APP-ABC-Program/1.0"
        }
        
        async with httpx.AsyncClient() as client:
            response = await client.get(url, params=params, headers=headers, timeout=10.0)
            
            if response.status_code == 200:
                data = response.json()
                
                if "display_name" in data:
                    address = data["display_name"]
                    addr_parts = data.get("address", {})
                    
                    return {
                        "success": True,
                        "address": address,
                        "formatted_address": address,
                        "components": {
                            "road": addr_parts.get("road", ""),
                            "suburb": addr_parts.get("suburb", ""),
                            "city": addr_parts.get("city", addr_parts.get("town", addr_parts.get("village", ""))),
                            "state": addr_parts.get("state", ""),
                            "postcode": addr_parts.get("postcode", ""),
                            "country": addr_parts.get("country", "")
                        },
                        "lat": lat,
                        "lng": lng,
                        "provider": "openstreetmap"
                    }
        
        return {
            "success": False,
            "address": f"Location: {lat:.6f}, {lng:.6f}",
            "lat": lat,
            "lng": lng,
            "error": "Could not find address for these coordinates"
        }
        
    except Exception as e:
        logger.error(f"Reverse geocoding error: {str(e)}")
        return {
            "success": False,
            "address": f"Location: {lat:.6f}, {lng:.6f}",
            "lat": lat,
            "lng": lng,
            "error": str(e)
        }

# ==================== GOOGLE DRIVE OAUTH ====================
from google_drive_oauth import (
    get_authorization_url, 
    exchange_code_for_credentials,
    test_drive_connection,
    upload_image_to_drive
)
from fastapi.responses import RedirectResponse

@api_router.get("/drive/connect")
async def connect_google_drive(current_user: dict = Depends(get_current_user)):
    """Initiate Google Drive OAuth flow - credentials will be stored for THIS user"""
    try:
        authorization_url, state = get_authorization_url()
        
        # Store state in database for verification - linked to current user
        await db.oauth_states.update_one(
            {"user_id": current_user["id"]},
            {"$set": {
                "user_id": current_user["id"],
                "state": state,
                "created_at": datetime.now(timezone.utc).isoformat()
            }},
            upsert=True
        )
        
        return {"authorization_url": authorization_url}
    except Exception as e:
        logger.error(f"Failed to initiate OAuth: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to initiate OAuth: {str(e)}")

@api_router.get("/drive/callback")
async def drive_oauth_callback(code: str, state: str = None):
    """Handle Google Drive OAuth callback - stores credentials for the specific user who initiated the flow"""
    try:
        # Exchange code for credentials
        credentials = exchange_code_for_credentials(code)
        
        # Find user by state - MUST have a valid state to link to user
        user_id = None
        user_email = None
        if state:
            oauth_state = await db.oauth_states.find_one({"state": state})
            if oauth_state:
                user_id = oauth_state["user_id"]
                # Get user email for logging
                user = await db.users.find_one({"id": user_id}, {"_id": 0, "email": 1})
                user_email = user.get("email") if user else None
                # Clean up used state
                await db.oauth_states.delete_one({"state": state})
        
        if not user_id:
            logger.error("OAuth callback: No valid state found, cannot link credentials to user")
            frontend_url = os.environ.get("FRONTEND_URL", "https://abc-program.preview.emergentagent.com")
            return RedirectResponse(url=f"{frontend_url}/settings?drive_error=Invalid OAuth state. Please try again.")
        
        # Store credentials in the USER's record directly (multi-user safe)
        await db.users.update_one(
            {"id": user_id},
            {"$set": {
                "google_drive_credentials": {
                    **credentials,
                    "connected_at": datetime.now(timezone.utc).isoformat()
                }
            }}
        )
        
        logger.info(f"Google Drive connected for user {user_id} ({user_email})")
        
        # Get frontend URL for redirect
        frontend_url = os.environ.get("FRONTEND_URL", "https://abc-program.preview.emergentagent.com")
        
        # Redirect to settings page with success message
        return RedirectResponse(url=f"{frontend_url}/settings?drive_connected=true")
        
    except Exception as e:
        logger.error(f"OAuth callback failed: {str(e)}")
        frontend_url = os.environ.get("FRONTEND_URL", "https://abc-program.preview.emergentagent.com")
        return RedirectResponse(url=f"{frontend_url}/settings?drive_error={str(e)}")

@api_router.get("/drive/status")
async def get_drive_status(current_user: dict = Depends(get_current_user)):
    """Check Google Drive connection status for the CURRENT user"""
    # Get user's own credentials from their user record
    user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0, "google_drive_credentials": 1})
    creds = user.get("google_drive_credentials") if user else None
    
    if not creds:
        return {
            "connected": False, 
            "message": "Your Google Drive is not connected. Click 'Connect Google Drive' to link your account."
        }
    
    # Test the connection
    result = test_drive_connection(creds)
    if result.get("connected"):
        result["connected_at"] = creds.get("connected_at")
    return result

@api_router.post("/drive/disconnect")
async def disconnect_google_drive(current_user: dict = Depends(get_current_user)):
    """Disconnect Google Drive for the current user"""
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$unset": {"google_drive_credentials": ""}}
    )
    logger.info(f"Google Drive disconnected for user {current_user['id']}")
    return {"message": "Google Drive disconnected successfully"}

@api_router.post("/drive/upload-test")
async def test_drive_upload(current_user: dict = Depends(get_current_user)):
    """Test upload a small image to Google Drive using current user's credentials"""
    # Get current user's credentials
    user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0, "google_drive_credentials": 1})
    creds = user.get("google_drive_credentials") if user else None
    
    if not creds:
        raise HTTPException(
            status_code=400, 
            detail="Your Google Drive is not connected. Please connect your account first."
        )
    
    # Test image (1x1 red pixel PNG)
    test_image_base64 = "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8DwHwAFBQIAX8jx0gAAAABJRU5ErkJggg=="
    
    folder_id = os.environ.get("GOOGLE_DRIVE_FOLDER_ID")
    
    result = upload_image_to_drive(
        creds_data=creds,
        base64_data=test_image_base64,
        filename=f"test_upload_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png",
        folder_id=folder_id
    )
    
    if result:
        # Update credentials if they were refreshed
        if result.get('updated_credentials'):
            await db.users.update_one(
                {"id": current_user["id"]},
                {"$set": {"google_drive_credentials": {
                    **creds,
                    **result['updated_credentials']
                }}}
            )
        return {"success": True, "file": result}
    else:
        raise HTTPException(status_code=500, detail="Upload failed")

# ==================== USER MANAGEMENT ====================

@api_router.post("/users", response_model=User)
async def create_user(
    user_data: UserCreate,
    current_user: dict = Depends(lambda creds: require_roles([UserRole.SUPER_ADMIN, UserRole.ADMIN])(creds))
):
    """Create a new user"""
    from utils import generate_password
    
    # Check if email already exists
    existing = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # For non-Super Admin, users must have a project_id
    project_id = user_data.project_id if hasattr(user_data, 'project_id') else current_user.get("project_id")
    
    # Generate password
    password = generate_password(user_data.first_name, user_data.mobile)
    
    # Create user
    user_dict = user_data.model_dump()
    user_dict["id"] = str(uuid.uuid4())
    user_dict["project_id"] = project_id
    user_dict["password_hash"] = hash_password(password)
    user_dict["is_active"] = True
    user_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.users.insert_one(user_dict)
    logger.info(f"User created: {user_data.email}, Password: {password}")
    
    user_dict.pop("password_hash")
    user_dict['created_at'] = datetime.fromisoformat(user_dict['created_at'])
    return User(**user_dict)

@api_router.get("/users", response_model=List[User])
async def get_users(
    project_id: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Get users - filtered by project for non-Super Admin"""
    query = {}
    
    # Super Admin can see all users or filter by project
    if current_user.get("role") == UserRole.SUPER_ADMIN.value:
        if project_id:
            query["project_id"] = project_id
    else:
        # Non-Super Admin can only see users from their project
        query["project_id"] = current_user.get("project_id")
    
    users = await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(None)
    for user in users:
        if isinstance(user.get('created_at'), str):
            user['created_at'] = datetime.fromisoformat(user['created_at'])
    return users

# ==================== PROJECT MANAGEMENT ====================

class ProjectCreateRequest(BaseModel):
    """Request model for creating a new project"""
    organization_name: str = "Janice Smith Animal Welfare Trust"
    organization_shortcode: str = "JS"
    organization_logo_base64: Optional[str] = None
    project_name: str
    project_code: str  # 3 letters
    project_logo_base64: Optional[str] = None
    project_address: Optional[str] = None
    max_kennels: int = 300
    
    # Admin user for this project
    admin_first_name: str
    admin_last_name: str
    admin_email: EmailStr
    admin_mobile: str
    admin_password: str

@api_router.get("/projects")
async def get_projects(
    current_user: dict = Depends(get_current_user)
):
    """Get all projects - Super Admin only"""
    if current_user.get("role") != UserRole.SUPER_ADMIN.value:
        # Non-Super Admin can only see their own project
        project_id = current_user.get("project_id")
        if project_id:
            project = await db.projects.find_one({"id": project_id}, {"_id": 0})
            return [project] if project else []
        return []
    
    projects = await db.projects.find({}, {"_id": 0}).to_list(None)
    return projects

@api_router.get("/projects/{project_code}")
async def get_project_by_code(
    project_code: str,
    current_user: dict = Depends(get_current_user)
):
    """Get project by code"""
    project = await db.projects.find_one(
        {"project_code": project_code.upper()}, 
        {"_id": 0}
    )
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Check access
    if current_user.get("role") != UserRole.SUPER_ADMIN.value:
        if current_user.get("project_id") != project["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
    
    return project

@api_router.post("/projects")
async def create_project(
    project_data: ProjectCreateRequest,
    current_user: dict = Depends(get_current_user)
):
    """Create a new project - Super Admin only"""
    if current_user.get("role") != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="Only Super Admin can create projects")
    
    # Validate project code (3 letters)
    project_code = project_data.project_code.upper().strip()
    if len(project_code) != 3 or not project_code.isalpha():
        raise HTTPException(status_code=400, detail="Project code must be exactly 3 letters")
    
    # Check if project code already exists
    existing = await db.projects.find_one({"project_code": project_code})
    if existing:
        raise HTTPException(status_code=400, detail=f"Project code '{project_code}' already exists")
    
    # Check if admin email already exists
    existing_user = await db.users.find_one({"email": project_data.admin_email})
    if existing_user:
        raise HTTPException(status_code=400, detail=f"Email '{project_data.admin_email}' already registered")
    
    # Create project
    project_id = str(uuid.uuid4())
    project = {
        "id": project_id,
        "organization_name": project_data.organization_name,
        "organization_shortcode": project_data.organization_shortcode.upper(),
        "organization_logo_url": None,  # Will be updated if logo uploaded
        "project_name": project_data.project_name,
        "project_code": project_code,
        "project_logo_url": None,  # Will be updated if logo uploaded
        "project_address": project_data.project_address,
        "max_kennels": project_data.max_kennels,
        "status": ProjectStatus.ACTIVE.value,
        "drive_folder_id": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.projects.insert_one(project)
    logger.info(f"Project created: {project_code} - {project_data.project_name}")
    
    # Create admin user for this project
    admin_user = {
        "id": str(uuid.uuid4()),
        "email": project_data.admin_email,
        "first_name": project_data.admin_first_name,
        "last_name": project_data.admin_last_name,
        "mobile": project_data.admin_mobile,
        "role": UserRole.ADMIN.value,
        "project_id": project_id,
        "password_hash": hash_password(project_data.admin_password),
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(admin_user)
    logger.info(f"Project admin created: {project_data.admin_email}")
    
    # Initialize kennels for this project
    kennels = []
    for i in range(1, project_data.max_kennels + 1):
        kennels.append({
            "id": str(uuid.uuid4()),
            "project_id": project_id,
            "kennel_number": i,
            "is_occupied": False,
            "current_case_id": None,
            "last_updated": datetime.now(timezone.utc).isoformat()
        })
    
    if kennels:
        await db.kennels.insert_many(kennels)
        logger.info(f"Initialized {project_data.max_kennels} kennels for project {project_code}")
    
    # Copy default medicines to this project
    default_medicines = await db.medicines.find({"project_id": {"$exists": False}}, {"_id": 0}).to_list(None)
    if default_medicines:
        for med in default_medicines:
            med["id"] = str(uuid.uuid4())
            med["project_id"] = project_id
            med["current_stock"] = 0
        await db.medicines.insert_many(default_medicines)
        logger.info(f"Copied {len(default_medicines)} medicines to project {project_code}")
    
    # Copy default food items to this project
    default_food = await db.food_items.find({"project_id": {"$exists": False}}, {"_id": 0}).to_list(None)
    if default_food:
        for food in default_food:
            food["id"] = str(uuid.uuid4())
            food["project_id"] = project_id
            food["current_stock"] = 0
        await db.food_items.insert_many(default_food)
        logger.info(f"Copied {len(default_food)} food items to project {project_code}")
    
    return {
        "message": f"Project '{project_data.project_name}' created successfully",
        "project": project,
        "admin_email": project_data.admin_email,
        "project_url": f"/app/{project_code.lower()}/"
    }

@api_router.put("/projects/{project_id}")
async def update_project(
    project_id: str,
    update_data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Update project details"""
    # Check access
    if current_user.get("role") != UserRole.SUPER_ADMIN.value:
        if current_user.get("project_id") != project_id:
            raise HTTPException(status_code=403, detail="Access denied")
        # Non-Super Admin can only update limited fields
        allowed_fields = ["project_address", "max_kennels"]
        update_data = {k: v for k, v in update_data.items() if k in allowed_fields}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No valid fields to update")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.projects.update_one(
        {"id": project_id},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    
    return {"message": "Project updated successfully"}

@api_router.delete("/projects/{project_id}")
async def delete_project(
    project_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a project - Super Admin only (sets status to Inactive)"""
    if current_user.get("role") != UserRole.SUPER_ADMIN.value:
        raise HTTPException(status_code=403, detail="Only Super Admin can delete projects")
    
    # Soft delete - set status to Inactive
    result = await db.projects.update_one(
        {"id": project_id},
        {"$set": {"status": ProjectStatus.INACTIVE.value, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    
    return {"message": "Project deactivated successfully"}

# ==================== STATISTICS ====================

@api_router.get("/statistics/dashboard")
async def get_dashboard_statistics(
    project_id: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Get dashboard statistics - filtered by project"""
    # Build query based on project access
    query = {}
    kennel_query = {}
    
    if current_user.get("role") == UserRole.SUPER_ADMIN.value:
        # Super Admin can see all or filter by project
        if project_id:
            query["project_id"] = project_id
            kennel_query["project_id"] = project_id
    else:
        # Non-Super Admin can only see their project
        user_project_id = current_user.get("project_id")
        if user_project_id:
            query["project_id"] = user_project_id
            kennel_query["project_id"] = user_project_id
    
    total_cases = await db.cases.count_documents(query)
    
    active_query = {**query, "status": {"$nin": [
        CaseStatus.RELEASED.value,
        CaseStatus.DECEASED.value,
        CaseStatus.SURGERY_CANCELLED.value
    ]}}
    active_cases = await db.cases.count_documents(active_query)
    
    surgery_query = {**query, "surgery": {"$exists": True}}
    total_surgeries = await db.cases.count_documents(surgery_query)
    
    occupied_kennel_query = {**kennel_query, "is_occupied": True}
    occupied_kennels = await db.kennels.count_documents(occupied_kennel_query)
    total_kennels = await db.kennels.count_documents(kennel_query)
    
    return {
        "total_cases": total_cases,
        "active_cases": active_cases,
        "total_surgeries": total_surgeries,
        "occupied_kennels": occupied_kennels,
        "available_kennels": total_kennels - occupied_kennels,
        "total_kennels": total_kennels
    }

# ==================== MEDICINE MANAGEMENT ====================
from models import Medicine, MedicineCreate, MedicineStockAdd

@api_router.post("/medicines", response_model=Medicine)
async def create_medicine(
    medicine_data: MedicineCreate,
    project_id: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Create a new medicine for a project"""
    # Determine project_id
    if current_user.get("role") == UserRole.SUPER_ADMIN.value:
        if not project_id:
            raise HTTPException(status_code=400, detail="project_id required for Super Admin")
    else:
        project_id = current_user.get("project_id")
    
    medicine_dict = medicine_data.model_dump()
    medicine_dict["id"] = str(uuid.uuid4())
    medicine_dict["project_id"] = project_id
    medicine_dict["current_stock"] = 0.0
    medicine_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.medicines.insert_one(medicine_dict)
    medicine_dict['created_at'] = datetime.fromisoformat(medicine_dict['created_at'])
    return Medicine(**medicine_dict)

@api_router.get("/medicines", response_model=List[Medicine])
async def get_medicines(
    project_id: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Get medicines - filtered by project"""
    query = {}
    
    if current_user.get("role") == UserRole.SUPER_ADMIN.value:
        if project_id:
            query["project_id"] = project_id
    else:
        user_project_id = current_user.get("project_id")
        if user_project_id:
            query["project_id"] = user_project_id
    
    medicines = await db.medicines.find(query, {"_id": 0}).to_list(None)
    for med in medicines:
        if isinstance(med.get('created_at'), str):
            med['created_at'] = datetime.fromisoformat(med['created_at'])
    return medicines

@api_router.post("/medicines/stock/add")
async def add_medicine_stock(
    stock_data: MedicineStockAdd,
    current_user: dict = Depends(get_current_user)
):
    """Add medicine stock - quantity is in PACKS, converted to units based on packing_size"""
    # Get the medicine to find packing_size
    medicine = await db.medicines.find_one({"id": stock_data.medicine_id}, {"_id": 0})
    if not medicine:
        raise HTTPException(status_code=404, detail="Medicine not found")
    
    # Check project access
    if current_user.get("role") != UserRole.SUPER_ADMIN.value:
        if medicine.get("project_id") != current_user.get("project_id"):
            raise HTTPException(status_code=403, detail="Access denied")
    
    packing_size = medicine.get("packing_size", 1)
    
    # Convert packs to units
    units_to_add = stock_data.quantity * packing_size
    
    # Update stock
    result = await db.medicines.update_one(
        {"id": stock_data.medicine_id},
        {"$inc": {"current_stock": units_to_add}}
    )
    
    # Log the stock addition - ensure date has time component for consistent querying
    stock_date = stock_data.date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if "T" not in stock_date:
        stock_date = stock_date + "T00:00:00"
    
    log_entry = {
        "id": str(uuid.uuid4()),
        "medicine_id": stock_data.medicine_id,
        "medicine_name": medicine.get("name"),
        "type": "restock",
        "packs_added": stock_data.quantity,
        "units_added": units_to_add,
        "packing_size": packing_size,
        "batch_number": stock_data.batch_number,
        "expiry_date": stock_data.expiry_date,
        "date": stock_date,
        "user_id": current_user["id"],
        "user_name": f"{current_user.get('first_name', '')} {current_user.get('last_name', '')}",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.medicine_logs.insert_one(log_entry)
    
    return {
        "message": "Stock added successfully", 
        "packs_added": stock_data.quantity,
        "units_added": units_to_add,
        "packing_size": packing_size
    }

@api_router.get("/medicines/stock/reset")
async def reset_medicine_stock(
    current_user: dict = Depends(require_roles([UserRole.SUPER_USER, UserRole.ADMIN]))
):
    """Reset all medicine stock to 0 and clear logs (for fixing data issues)"""
    await db.medicines.update_many({}, {"$set": {"current_stock": 0}})
    await db.medicine_logs.delete_many({})
    return {"message": "All medicine stock reset to 0"}

@api_router.get("/medicines/logs")
async def get_medicine_logs(
    start_date: str = None,
    end_date: str = None,
    medicine_id: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Get medicine usage and restock logs"""
    from datetime import datetime as dt
    
    query = {}
    
    # Filter by user-provided 'date' field - convert strings to datetime for MongoDB
    if start_date:
        try:
            start_dt = dt.fromisoformat(start_date.replace(" ", "T") + ("T00:00:00" if "T" not in start_date and " " not in start_date else ""))
            query["date"] = {"$gte": start_dt}
        except ValueError:
            query["date"] = {"$gte": start_date}
    if end_date:
        try:
            end_str = end_date + ("T23:59:59" if "T" not in end_date and " " not in end_date else "")
            end_dt = dt.fromisoformat(end_str.replace(" ", "T"))
            if "date" in query:
                query["date"]["$lte"] = end_dt
            else:
                query["date"] = {"$lte": end_dt}
        except ValueError:
            if "date" in query:
                query["date"]["$lte"] = end_date
            else:
                query["date"] = {"$lte": end_date}
    if medicine_id:
        query["medicine_id"] = medicine_id
    
    logs = await db.medicine_logs.find(query, {"_id": 0}).sort("date", -1).to_list(None)
    
    # Convert datetime objects to ISO strings for JSON serialization
    for log in logs:
        if isinstance(log.get("date"), datetime):
            log["date"] = log["date"].isoformat()
        if isinstance(log.get("created_at"), datetime):
            log["created_at"] = log["created_at"].isoformat()
    
    return logs

@api_router.get("/medicines/usage-report")
async def get_medicine_usage_report(
    period: str = "month",
    month: str = None,
    week: int = None,
    start_date: str = None,
    end_date: str = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Get medicine usage report for a period.
    period: 'month', 'week', 'custom'
    month: YYYY-MM format (for month/week period)
    week: 1-5 (for week period)
    start_date, end_date: YYYY-MM-DD (for custom period)
    """
    from datetime import date
    import calendar
    
    now = datetime.now(timezone.utc)
    
    if period == "month":
        if month:
            year, mon = map(int, month.split('-'))
        else:
            year, mon = now.year, now.month
        
        first_day = date(year, mon, 1)
        last_day = date(year, mon, calendar.monthrange(year, mon)[1])
        start = first_day.isoformat() + "T00:00:00"
        end = last_day.isoformat() + "T23:59:59"
        period_name = f"{calendar.month_name[mon]} {year}"
        
    elif period == "week":
        if month:
            year, mon = map(int, month.split('-'))
        else:
            year, mon = now.year, now.month
        
        week_num = week or 1
        first_day_of_month = date(year, mon, 1)
        
        # Calculate week boundaries
        # Week 1: days 1-7, Week 2: days 8-14, etc.
        week_start_day = (week_num - 1) * 7 + 1
        week_end_day = min(week_num * 7, calendar.monthrange(year, mon)[1])
        
        start = date(year, mon, week_start_day).isoformat() + "T00:00:00"
        end = date(year, mon, week_end_day).isoformat() + "T23:59:59"
        period_name = f"Week {week_num} of {calendar.month_name[mon]} {year}"
        
    elif period == "custom":
        if not start_date or not end_date:
            raise HTTPException(status_code=400, detail="start_date and end_date required for custom period")
        start = start_date + "T00:00:00" if "T" not in start_date else start_date
        end = end_date + "T23:59:59" if "T" not in end_date else end_date
        period_name = f"{start_date} to {end_date}"
    else:
        raise HTTPException(status_code=400, detail="Invalid period. Use 'month', 'week', or 'custom'")
    
    # Get all logs in the period - filter by user-provided 'date' field
    # Convert string dates to datetime for proper MongoDB comparison
    from datetime import datetime as dt
    
    # Parse start and end dates - handle both T and space formats
    try:
        start_dt = dt.fromisoformat(start.replace(" ", "T"))
        end_dt = dt.fromisoformat(end.replace(" ", "T"))
    except ValueError:
        # Fallback if parsing fails
        start_dt = dt.strptime(start.replace("T", " "), "%Y-%m-%d %H:%M:%S")
        end_dt = dt.strptime(end.replace("T", " "), "%Y-%m-%d %H:%M:%S")
    
    logs = await db.medicine_logs.find({
        "date": {"$gte": start_dt, "$lte": end_dt}
    }, {"_id": 0}).to_list(None)
    
    # Get all medicines
    medicines = await db.medicines.find({}, {"_id": 0}).to_list(None)
    medicine_map = {m["id"]: m for m in medicines}
    
    # Aggregate by medicine
    usage_summary = {}
    for log in logs:
        med_id = log.get("medicine_id")
        if med_id not in usage_summary:
            med = medicine_map.get(med_id, {})
            usage_summary[med_id] = {
                "medicine_id": med_id,
                "medicine_name": log.get("medicine_name") or med.get("name", "Unknown"),
                "unit": med.get("unit", ""),
                "packing_size": med.get("packing_size", 1),
                "current_stock": med.get("current_stock", 0),
                "restocked_units": 0,
                "restocked_packs": 0,
                "used_units": 0,
                "restock_entries": [],
                "usage_entries": []
            }
        
        # Convert datetime to ISO string for JSON serialization
        log_date = log.get("date")
        if isinstance(log_date, datetime):
            log_date = log_date.isoformat()
        
        if log.get("type") == "restock":
            usage_summary[med_id]["restocked_units"] += log.get("units_added", 0)
            usage_summary[med_id]["restocked_packs"] += log.get("packs_added", 0)
            usage_summary[med_id]["restock_entries"].append({
                "date": log_date,
                "packs": log.get("packs_added", 0),
                "units": log.get("units_added", 0),
                "batch": log.get("batch_number"),
                "user": log.get("user_name")
            })
        elif log.get("type") == "usage":
            created_at = log.get("created_at")
            if isinstance(created_at, datetime):
                created_at = created_at.isoformat()
            usage_summary[med_id]["used_units"] += log.get("units_used", 0)
            usage_summary[med_id]["usage_entries"].append({
                "date": created_at,
                "case_number": log.get("case_number"),
                "units": log.get("units_used", 0),
                "user": log.get("user_name")
            })
    
    return {
        "period": period_name,
        "start_date": start,
        "end_date": end,
        "summary": list(usage_summary.values()),
        "total_restock_entries": len([l for l in logs if l.get("type") == "restock"]),
        "total_usage_entries": len([l for l in logs if l.get("type") == "usage"])
    }

@api_router.put("/medicines/{medicine_id}")
async def update_medicine(
    medicine_id: str,
    medicine_data: dict,
    current_user: dict = Depends(require_roles([UserRole.SUPER_USER, UserRole.ADMIN]))
):
    """Update medicine details (name, generic_name, unit, packing, packing_size)"""
    # Fields that can be updated
    allowed_fields = ["name", "generic_name", "unit", "packing", "packing_size"]
    update_data = {k: v for k, v in medicine_data.items() if k in allowed_fields}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No valid fields to update")
    
    # Add updated timestamp
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.medicines.update_one(
        {"id": medicine_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Medicine not found")
    
    return {"message": "Medicine updated successfully"}

@api_router.delete("/medicines/{medicine_id}")
async def delete_medicine(
    medicine_id: str,
    current_user: dict = Depends(require_roles([UserRole.SUPER_USER, UserRole.ADMIN]))
):
    """Delete a medicine"""
    result = await db.medicines.delete_one({"id": medicine_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Medicine not found")
    
    return {"message": "Medicine deleted successfully"}

# ==================== FOOD MANAGEMENT ====================
from models import FoodItem, FoodItemCreate, FoodStockAdd

@api_router.post("/food-items", response_model=FoodItem)
async def create_food_item(
    food_data: FoodItemCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a new food item"""
    food_dict = food_data.model_dump()
    food_dict["id"] = str(uuid.uuid4())
    food_dict["current_stock"] = 0.0
    food_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.food_items.insert_one(food_dict)
    food_dict['created_at'] = datetime.fromisoformat(food_dict['created_at'])
    return FoodItem(**food_dict)

@api_router.get("/food-items", response_model=List[FoodItem])
async def get_food_items(current_user: dict = Depends(get_current_user)):
    """Get all food items"""
    items = await db.food_items.find({}, {"_id": 0}).to_list(None)
    for item in items:
        if isinstance(item.get('created_at'), str):
            item['created_at'] = datetime.fromisoformat(item['created_at'])
    return items

@api_router.post("/food-items/stock/add")
async def add_food_stock(
    stock_data: FoodStockAdd,
    current_user: dict = Depends(get_current_user)
):
    """Add food stock"""
    result = await db.food_items.update_one(
        {"id": stock_data.food_id},
        {"$inc": {"current_stock": stock_data.quantity}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Food item not found")
    
    return {"message": "Stock added successfully", "quantity": stock_data.quantity}

# ==================== KENNEL MANAGEMENT ====================

@api_router.get("/kennels")
async def get_kennels(
    status_filter: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all kennels or filter by status"""
    query = {}
    if status_filter == "available":
        query["is_occupied"] = False
    elif status_filter == "occupied":
        query["is_occupied"] = True
    
    kennels = await db.kennels.find(query, {"_id": 0}).sort("kennel_number", 1).to_list(None)
    return kennels

@api_router.post("/kennels/initialize")
async def initialize_kennels(current_user: dict = Depends(get_current_user)):
    """Initialize kennels (run once)"""
    config = await db.system_config.find_one({"id": "system_config"}, {"_id": 0})
    max_kennels = config.get("max_kennels", 300) if config else 300
    
    existing = await db.kennels.count_documents({})
    if existing > 0:
        return {"message": f"{existing} kennels already exist"}
    
    kennels = []
    for i in range(1, max_kennels + 1):
        kennel = {
            "id": str(uuid.uuid4()),
            "kennel_number": i,
            "is_occupied": False,
            "current_case_id": None,
            "last_updated": datetime.now(timezone.utc).isoformat()
        }
        kennels.append(kennel)
    
    if kennels:
        await db.kennels.insert_many(kennels)
    
    return {"message": f"Initialized {len(kennels)} kennels"}

# ==================== CASE MANAGEMENT ====================
from drive_uploader import get_drive_uploader, get_drive_uploader_for_user

@api_router.post("/cases/catching")
async def create_catching_record(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Create a catching record with case number format: JS-TAL-JAN-C0001"""
    from utils import get_next_case_number
    
    config = await db.system_config.find_one({"id": "system_config"}, {"_id": 0})
    org_shortcode = config.get("organization_shortcode", "JS") if config else "JS"
    project_code = config.get("project_code", "TAL") if config else "TAL"
    # Generate case number with "C" prefix for catching
    case_number = await get_next_case_number(db, org_shortcode, project_code, case_type="C")
    
    # Upload photos to Google Drive (using current user's credentials)
    photo_links = []
    drive_uploader = await get_drive_uploader_for_user(db, current_user)
    
    if drive_uploader:
        photos = data.get("photos", [])
        if not photos and data.get("photo_base64"):
            photos = [data["photo_base64"]]
        
        catching_date = datetime.fromisoformat(data.get("date_time", datetime.now(timezone.utc).isoformat()).replace('Z', '+00:00')) if data.get("date_time") else datetime.now(timezone.utc)
        
        for i, photo in enumerate(photos[:4]):
            if photo:
                result = drive_uploader.upload_image(
                    base64_data=photo,
                    form_type="Catching",
                    case_number=case_number,
                    date=catching_date,
                    photo_index=i
                )
                if result:
                    photo_links.append(result)
        
        # Update user's credentials if refreshed
        updated_creds = drive_uploader.get_updated_credentials()
        if updated_creds.get("access_token") != drive_uploader.creds_data.get("access_token"):
            await db.users.update_one(
                {"id": current_user["id"]},
                {"$set": {"google_drive_credentials": updated_creds}}
            )
    
    case_dict = {
        "id": str(uuid.uuid4()),
        "case_number": case_number,
        "status": CaseStatus.CAUGHT.value,
        "project_code": project_code,
        "catching": {
            "date_time": data.get("date_time", datetime.now(timezone.utc).isoformat()),
            "location_lat": data["location_lat"],
            "location_lng": data["location_lng"],
            "address": data["address"],
            "ward_number": data.get("ward_number"),
            "photo_links": photo_links,  # Store Google Drive links instead of base64
            "photo_base64": data.get("photo_base64") if not photo_links else None,  # Fallback if Drive fails
            "remarks": data.get("remarks"),
            "driver_id": current_user["id"]
        },
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.cases.insert_one(case_dict)
    logger.info(f"Case created: {case_number} with {len(photo_links)} photos uploaded to Drive")
    
    return {"case_number": case_number, "case_id": case_dict["id"], "message": "Case created successfully", "photos_uploaded": len(photo_links)}

@api_router.get("/cases")
async def get_cases(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all cases"""
    query = {}
    if status:
        query["status"] = status
    
    cases = await db.cases.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return cases

@api_router.get("/cases/{case_id}")
async def get_case(
    case_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get a specific case"""
    case = await db.cases.find_one({"id": case_id}, {"_id": 0})
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    return case

@api_router.post("/cases/{case_id}/initial-observation")
async def add_initial_observation(
    case_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Add initial observation to a case"""
    case = await db.cases.find_one({"id": case_id}, {"_id": 0})
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    
    # Check if kennel is available
    kennel = await db.kennels.find_one({"kennel_number": data["kennel_number"]}, {"_id": 0})
    if not kennel or kennel["is_occupied"]:
        raise HTTPException(status_code=400, detail="Kennel not available")
    
    # Update case
    observation = {
        "kennel_number": data["kennel_number"],
        "gender": data["gender"],
        "approximate_age": data["approximate_age"],
        "color_markings": data["color_markings"],
        "body_condition": data["body_condition"],
        "temperament": data["temperament"],
        "visible_injuries": data["visible_injuries"],
        "injury_description": data.get("injury_description"),
        "photo_base64": data["photo_base64"],
        "remarks": data.get("remarks"),
        "catcher_id": current_user["id"],
        "observation_date": datetime.now(timezone.utc).isoformat()
    }
    
    await db.cases.update_one(
        {"id": case_id},
        {
            "$set": {
                "initial_observation": observation,
                "status": CaseStatus.IN_KENNEL.value,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    # Update kennel
    await db.kennels.update_one(
        {"kennel_number": data["kennel_number"]},
        {
            "$set": {
                "is_occupied": True,
                "current_case_id": case_id,
                "last_updated": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    return {"message": "Initial observation added successfully"}

@api_router.post("/cases/{case_id}/surgery")
async def add_surgery_record(
    case_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Add surgery record to a case"""
    case = await db.cases.find_one({"id": case_id}, {"_id": 0})
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    
    # Upload photos to Google Drive (using current user's credentials)
    photo_links = []
    drive_uploader = await get_drive_uploader_for_user(db, current_user)
    
    if drive_uploader and data.get("photos"):
        surgery_date = datetime.fromisoformat(data.get("surgery_date", datetime.now(timezone.utc).isoformat()).replace('Z', '+00:00')) if data.get("surgery_date") else datetime.now(timezone.utc)
        
        for i, photo in enumerate(data.get("photos", [])[:4]):
            if photo:
                result = drive_uploader.upload_image(
                    base64_data=photo,
                    form_type="Surgery",
                    case_number=case["case_number"],
                    date=surgery_date,
                    photo_index=i
                )
                if result:
                    photo_links.append(result)
        
        # Update user's credentials if refreshed
        updated_creds = drive_uploader.get_updated_credentials()
        if updated_creds.get("access_token") != drive_uploader.creds_data.get("access_token"):
            await db.users.update_one(
                {"id": current_user["id"]},
                {"$set": {"google_drive_credentials": updated_creds}}
            )
    
    # Get weight and gender for auto medicine calculation
    weight = data.get("weight", 0)
    gender = data.get("gender") or (case.get("initial_observation", {}).get("gender"))
    
    # Auto-calculate medicines if weight provided and surgery not cancelled
    medicines_used = data.get("medicines", {})
    medicines_to_deduct = data.get("medicines_to_deduct", {})
    
    # If weight provided but no explicit medicines_to_deduct, auto-calculate
    if weight and weight >= 10 and weight <= 30 and data.get("pre_surgery_status") != "Cancel Surgery":
        if not medicines_to_deduct:
            # Get all medicines for mapping
            all_medicines = await db.medicines.find({}, {"_id": 0}).to_list(None)
            medicine_map = {m["name"]: m for m in all_medicines}
            
            for med_name in MEDICINE_PROTOCOL.keys():
                dosage = calculate_medicine_dosage(weight, med_name, gender)
                if dosage > 0 and med_name in medicine_map:
                    med_id = medicine_map[med_name]["id"]
                    medicines_to_deduct[med_id] = dosage
                    medicines_used[med_name] = dosage
    
    surgery = {
        "surgery_date": data.get("surgery_date", datetime.now(timezone.utc).isoformat()),
        "pre_surgery_status": data.get("pre_surgery_status", "Fit for Surgery") if not data.get("cancelled") or data.get("cancelled") == "No" else "Cancel Surgery",
        "cancellation_reason": data.get("cancellation_reason"),
        "surgery_type": data.get("surgery_type") or ("Ovariohysterectomy" if gender == "Female" else "Castration"),
        "weight": weight,
        "skin": data.get("skin", "Normal"),
        "anesthesia_used": data.get("anesthesia_used", []),
        "surgery_start_time": data.get("surgery_start_time"),
        "surgery_end_time": data.get("surgery_end_time"),
        "complications": data.get("complications", False),
        "complication_description": data.get("complication_description"),
        "post_surgery_status": data.get("post_surgery_status", "Good"),
        "veterinary_signature": data.get("veterinary_signature", f"{current_user.get('first_name', '')} {current_user.get('last_name', '')}"),
        "remarks": data.get("remarks"),
        "veterinary_id": current_user["id"],
        "photo_links": photo_links,
        "medicines_used": medicines_used
    }
    
    # Determine new status
    is_cancelled = data.get("cancelled") == "Yes" or data.get("pre_surgery_status") == "Cancel Surgery"
    if is_cancelled:
        surgery["pre_surgery_status"] = "Cancel Surgery"
        new_status = CaseStatus.SURGERY_CANCELLED.value
        # Free kennel if cancelled
        if case.get("initial_observation"):
            kennel_number = case["initial_observation"].get("kennel_number")
            if kennel_number:
                await db.kennels.update_one(
                    {"kennel_number": kennel_number},
                    {
                        "$set": {
                            "is_occupied": False,
                            "current_case_id": None,
                            "last_updated": datetime.now(timezone.utc).isoformat()
                        }
                    }
                )
    else:
        new_status = CaseStatus.SURGERY_COMPLETED.value
        # Deduct medicine stock and log usage
        for medicine_id, quantity in medicines_to_deduct.items():
            if quantity > 0:
                # Get medicine details
                medicine = await db.medicines.find_one({"id": medicine_id}, {"_id": 0})
                
                # Deduct from stock
                await db.medicines.update_one(
                    {"id": medicine_id},
                    {"$inc": {"current_stock": -quantity}}
                )
                
                # Log the usage
                log_entry = {
                    "id": str(uuid.uuid4()),
                    "medicine_id": medicine_id,
                    "medicine_name": medicine.get("name") if medicine else "Unknown",
                    "type": "usage",
                    "units_used": quantity,
                    "case_id": case_id,
                    "case_number": case.get("case_number"),
                    "user_id": current_user["id"],
                    "user_name": f"{current_user.get('first_name', '')} {current_user.get('last_name', '')}",
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.medicine_logs.insert_one(log_entry)
    
    await db.cases.update_one(
        {"id": case_id},
        {
            "$set": {
                "surgery": surgery,
                "status": new_status,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    return {"message": "Surgery record added successfully", "status": new_status, "photos_uploaded": len(photo_links)}

@api_router.post("/cases/{case_id}/treatment")
async def add_daily_treatment(
    case_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Add daily treatment record"""
    case = await db.cases.find_one({"id": case_id}, {"_id": 0})
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    
    # Upload photos to Google Drive (Post-op-care) using current user's credentials
    photo_links = []
    drive_uploader = await get_drive_uploader_for_user(db, current_user)
    
    if drive_uploader and data.get("photos"):
        treatment_date = datetime.fromisoformat(data.get("date", datetime.now(timezone.utc).isoformat()).replace('Z', '+00:00')) if data.get("date") else datetime.now(timezone.utc)
        
        for i, photo in enumerate(data.get("photos", [])[:4]):
            if photo:
                result = drive_uploader.upload_image(
                    base64_data=photo,
                    form_type="Post-op-care",
                    case_number=case["case_number"],
                    date=treatment_date,
                    photo_index=i
                )
                if result:
                    photo_links.append(result)
        
        # Update user's credentials if refreshed
        updated_creds = drive_uploader.get_updated_credentials()
        if updated_creds.get("access_token") != drive_uploader.creds_data.get("access_token"):
            await db.users.update_one(
                {"id": current_user["id"]},
                {"$set": {"google_drive_credentials": updated_creds}}
            )
    
    treatment = {
        "id": str(uuid.uuid4()),
        "case_id": case_id,
        "date": data.get("date", datetime.now(timezone.utc).isoformat()),
        "day_post_surgery": data["day_post_surgery"],
        # Support both old format (medicine IDs) and new format (medicines_used dict)
        "medicines_used": data.get("medicines_used", {}),
        "antibiotic_id": data.get("antibiotic_id"),
        "antibiotic_dosage": data.get("antibiotic_dosage"),
        "painkiller_id": data.get("painkiller_id"),
        "painkiller_dosage": data.get("painkiller_dosage"),
        "additional_medicine_id": data.get("additional_medicine_id"),
        "additional_medicine_dosage": data.get("additional_medicine_dosage"),
        "wound_condition": data["wound_condition"],
        "food_intake": data.get("food_intake", True),
        "water_intake": data.get("water_intake", True),
        "remarks": data.get("remarks"),
        "admin_id": current_user["id"],
        "photo_links": photo_links
    }
    
    # Deduct medicine stock - handle both old and new formats
    if data.get("medicines_used"):
        # New format: medicines_used dict with medicine names and dosages
        # Note: This doesn't deduct from stock since we're tracking by name not ID
        # For proper stock management, medicines should be matched by name to their IDs
        pass
    else:
        # Old format with medicine IDs
        medicines_to_deduct = [
            (data.get("antibiotic_id"), data.get("antibiotic_dosage")),
            (data.get("painkiller_id"), data.get("painkiller_dosage")),
            (data.get("additional_medicine_id"), data.get("additional_medicine_dosage"))
        ]
        
        for med_id, dosage in medicines_to_deduct:
            if med_id and dosage:
                await db.medicines.update_one(
                    {"id": med_id},
                    {"$inc": {"current_stock": -float(dosage)}}
                )
    
    # Add treatment to case
    await db.cases.update_one(
        {"id": case_id},
        {
            "$push": {"daily_treatments": treatment},
            "$set": {
                "status": CaseStatus.UNDER_TREATMENT.value,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    return {"message": "Treatment record added successfully", "photos_uploaded": len(photo_links)}

@api_router.post("/daily-feeding")
async def create_daily_feeding(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Create daily feeding record"""
    
    # Upload photos to Google Drive (Feeding) using current user's credentials
    photo_links = []
    drive_uploader = await get_drive_uploader_for_user(db, current_user)
    
    if drive_uploader and data.get("photos"):
        feeding_date = datetime.fromisoformat(data.get("date", datetime.now(timezone.utc).isoformat()).replace('Z', '+00:00')) if data.get("date") else datetime.now(timezone.utc)
        
        # For feeding, we use a combined identifier since it's not case-specific
        feeding_id = f"FEED-{feeding_date.strftime('%d%m%y')}-{data['meal_time']}"
        
        for i, photo in enumerate(data.get("photos", [])[:4]):
            if photo:
                result = drive_uploader.upload_image(
                    base64_data=photo,
                    form_type="Feeding",
                    case_number=feeding_id,
                    date=feeding_date,
                    photo_index=i
                )
                if result:
                    photo_links.append(result)
        
        # Update user's credentials if refreshed
        updated_creds = drive_uploader.get_updated_credentials()
        if updated_creds.get("access_token") != drive_uploader.creds_data.get("access_token"):
            await db.users.update_one(
                {"id": current_user["id"]},
                {"$set": {"google_drive_credentials": updated_creds}}
            )
    
    feeding = {
        "id": str(uuid.uuid4()),
        "date": data.get("date", datetime.now(timezone.utc).isoformat()),
        "meal_time": data["meal_time"],
        "kennel_numbers": data["kennel_numbers"],
        "food_items": data["food_items"],
        "total_quantity": data["total_quantity"],
        "photo_links": photo_links,
        "photo_base64": data.get("photo_base64") if not photo_links else None,
        "animals_not_fed": data.get("animals_not_fed", []),
        "remarks": data.get("remarks"),
        "caretaker_id": current_user["id"]
    }
    
    # Deduct food stock
    for food_id, quantity in data["food_items"].items():
        await db.food_items.update_one(
            {"id": food_id},
            {"$inc": {"current_stock": -float(quantity)}}
        )
    
    await db.daily_feeding.insert_one(feeding)
    
    # Update cases with feeding reference
    for kennel_num in data["kennel_numbers"]:
        kennel = await db.kennels.find_one({"kennel_number": kennel_num}, {"_id": 0})
        if kennel and kennel.get("current_case_id"):
            await db.cases.update_one(
                {"id": kennel["current_case_id"]},
                {"$push": {"daily_feedings": feeding["id"]}}
            )
    
    return {"message": "Feeding record created successfully", "feeding_id": feeding["id"], "photos_uploaded": len(photo_links)}

@api_router.post("/cases/{case_id}/release")
async def add_release_record(
    case_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Add release record"""
    case = await db.cases.find_one({"id": case_id}, {"_id": 0})
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    
    # Upload photos to Google Drive (Release) using current user's credentials
    photo_links = []
    drive_uploader = await get_drive_uploader_for_user(db, current_user)
    
    if drive_uploader:
        photos = data.get("photos", [])
        if not photos and data.get("photo_base64"):
            photos = [data["photo_base64"]]
        
        release_date = datetime.fromisoformat(data.get("date_time", datetime.now(timezone.utc).isoformat()).replace('Z', '+00:00')) if data.get("date_time") else datetime.now(timezone.utc)
        
        for i, photo in enumerate(photos[:4]):
            if photo:
                result = drive_uploader.upload_image(
                    base64_data=photo,
                    form_type="Release",
                    case_number=case["case_number"],
                    date=release_date,
                    photo_index=i
                )
                if result:
                    photo_links.append(result)
        
        # Update user's credentials if refreshed
        updated_creds = drive_uploader.get_updated_credentials()
        if updated_creds.get("access_token") != drive_uploader.creds_data.get("access_token"):
            await db.users.update_one(
                {"id": current_user["id"]},
                {"$set": {"google_drive_credentials": updated_creds}}
            )
    
    release = {
        "date_time": data.get("date_time", datetime.now(timezone.utc).isoformat()),
        "location_lat": data["location_lat"],
        "location_lng": data["location_lng"],
        "address": data["address"],
        "photo_links": photo_links,
        "photo_base64": data.get("photo_base64") if not photo_links else None,
        "released_by": current_user["id"],
        "remarks": data.get("remarks")
    }
    
    await db.cases.update_one(
        {"id": case_id},
        {
            "$set": {
                "release": release,
                "status": CaseStatus.RELEASED.value,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    # Free kennel
    if case.get("initial_observation"):
        kennel_number = case["initial_observation"]["kennel_number"]
        await db.kennels.update_one(
            {"kennel_number": kennel_number},
            {
                "$set": {
                    "is_occupied": False,
                    "current_case_id": None,
                    "last_updated": datetime.now(timezone.utc).isoformat()
                }
            }
        )
    
    return {"message": "Release record added successfully", "photos_uploaded": len(photo_links)}

# ==================== BULK UPLOAD ENDPOINTS ====================

@api_router.get("/bulk-upload/template/{template_type}")
async def download_bulk_upload_template(
    template_type: str,
    current_user: dict = Depends(get_current_user)
):
    """Download Excel template for bulk upload"""
    if template_type not in ["catching", "surgery"]:
        raise HTTPException(status_code=400, detail="Invalid template type. Use 'catching' or 'surgery'")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    required_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if template_type == "catching":
        ws.title = "Catching Records"
        headers = [
            ("Case Number*", "E.g., JAPP-001-2024", True),
            ("Date (DD/MM/YYYY)*", "E.g., 25/12/2024", True),
            ("Time (HH:MM)*", "E.g., 14:30", True),
            ("Latitude*", "E.g., 19.0760", True),
            ("Longitude*", "E.g., 72.8777", True),
            ("Address*", "Full address", True),
            ("Ward Number", "Optional", False),
            ("Remarks", "Optional notes", False),
        ]
    else:  # surgery
        ws.title = "Surgery Records"
        headers = [
            ("Case Number*", "E.g., JAPP-001-2024", True),
            ("Surgery Date (DD/MM/YYYY)*", "E.g., 26/12/2024", True),
            ("Gender*", "Male or Female", True),
            ("Weight (kg)*", "E.g., 15 (10-30 kg)", True),
            ("Surgery Cancelled*", "Yes or No", True),
            ("Cancellation Reason", "Required if cancelled (Too weak, Under age, Looks ill, etc.)", False),
            ("Skin Condition", "Normal, Rough, or Visible Infection", False),
            ("Remarks", "Optional notes", False),
        ]
    
    # Write headers
    for col, (header, hint, required) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        
        # Add hint row
        hint_cell = ws.cell(row=2, column=col, value=hint)
        hint_cell.font = Font(italic=True, color="666666")
        hint_cell.border = thin_border
        if required:
            hint_cell.fill = required_fill
        
        # Adjust column width
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(len(header), len(hint)) + 5
    
    # Add sample data row
    if template_type == "catching":
        sample_data = ["JAPP-001-2024", "25/12/2024", "14:30", "19.0760", "72.8777", "123 Main Street, Mumbai", "Ward 5", "Sample remark"]
    else:
        sample_data = ["JAPP-001-2024", "26/12/2024", "Male", "15", "No", "", "Normal", "Sample remark"]
    
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=3, column=col, value=value)
        cell.border = thin_border
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"bulk_upload_{template_type}_template.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.post("/bulk-upload/catching")
async def bulk_upload_catching(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Bulk upload catching records from Excel file"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents), header=0)  # Read with header row
        
        # Skip the hint row (second row, index 0 after header)
        if len(df) > 0 and str(df.iloc[0, 0]).startswith('E.g'):
            df = df.iloc[1:]  # Skip hint row
            df = df.reset_index(drop=True)
        
        # Normalize column names
        df.columns = df.columns.str.strip().str.lower().str.replace('*', '', regex=False)
        
        results = {"success": 0, "failed": 0, "errors": []}
        
        # Get project codes
        config = await db.system_config.find_one({"id": "system_config"}, {"_id": 0})
        org_shortcode = config.get("organization_shortcode", "JS") if config else "JS"
        project_code = config.get("project_code", "TAL") if config else "TAL"
        
        for idx, row in df.iterrows():
            row_num = idx + 3  # Account for header and hint rows
            try:
                # Validate required fields
                case_number = str(row.get('case number', '')).strip()
                if not case_number or case_number == 'nan':
                    results["errors"].append(f"Row {row_num}: Missing case number")
                    results["failed"] += 1
                    continue
                
                # Check if case already exists
                existing = await db.cases.find_one({"case_number": case_number})
                if existing:
                    results["errors"].append(f"Row {row_num}: Case {case_number} already exists")
                    results["failed"] += 1
                    continue
                
                # Parse date and time
                date_str = str(row.get('date (dd/mm/yyyy)', '')).strip()
                time_str = str(row.get('time (hh:mm)', '')).strip()
                
                if not date_str or date_str == 'nan':
                    results["errors"].append(f"Row {row_num}: Missing date")
                    results["failed"] += 1
                    continue
                
                # Parse the date
                try:
                    if isinstance(row.get('date (dd/mm/yyyy)'), datetime):
                        date_obj = row.get('date (dd/mm/yyyy)')
                    else:
                        date_obj = datetime.strptime(date_str, "%d/%m/%Y")
                except (ValueError, TypeError):
                    try:
                        date_obj = datetime.strptime(date_str.split()[0], "%Y-%m-%d")
                    except (ValueError, TypeError):
                        results["errors"].append(f"Row {row_num}: Invalid date format. Use DD/MM/YYYY")
                        results["failed"] += 1
                        continue
                
                # Parse time
                if time_str and time_str != 'nan':
                    try:
                        time_parts = time_str.split(':')
                        date_obj = date_obj.replace(hour=int(time_parts[0]), minute=int(time_parts[1]))
                    except (ValueError, IndexError):
                        pass
                
                date_obj = date_obj.replace(tzinfo=timezone.utc)
                
                # Validate coordinates
                try:
                    lat = float(row.get('latitude', 0))
                    lng = float(row.get('longitude', 0))
                    if lat == 0 or lng == 0:
                        results["errors"].append(f"Row {row_num}: Invalid coordinates")
                        results["failed"] += 1
                        continue
                except (ValueError, TypeError):
                    results["errors"].append(f"Row {row_num}: Invalid coordinates format")
                    results["failed"] += 1
                    continue
                
                address = str(row.get('address', '')).strip()
                if not address or address == 'nan':
                    results["errors"].append(f"Row {row_num}: Missing address")
                    results["failed"] += 1
                    continue
                
                ward_number = str(row.get('ward number', '')).strip()
                if ward_number == 'nan':
                    ward_number = None
                
                remarks = str(row.get('remarks', '')).strip()
                if remarks == 'nan':
                    remarks = None
                
                # Create case
                case_dict = {
                    "id": str(uuid.uuid4()),
                    "case_number": case_number,
                    "status": CaseStatus.CAUGHT.value,
                    "project_code": project_code,
                    "catching": {
                        "date_time": date_obj.isoformat(),
                        "location_lat": lat,
                        "location_lng": lng,
                        "address": address,
                        "ward_number": ward_number,
                        "photo_links": [],
                        "remarks": remarks,
                        "driver_id": current_user["id"],
                        "bulk_uploaded": True
                    },
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.cases.insert_one(case_dict)
                results["success"] += 1
                
            except Exception as e:
                results["errors"].append(f"Row {row_num}: {str(e)}")
                results["failed"] += 1
        
        return {
            "message": f"Bulk upload completed. {results['success']} records created, {results['failed']} failed.",
            "results": results
        }
        
    except Exception as e:
        logger.error(f"Bulk upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@api_router.post("/bulk-upload/surgery")
async def bulk_upload_surgery(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Bulk upload surgery records from Excel file"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents), header=0)  # Read with header row
        
        # Skip the hint row (second row, index 0 after header)
        if len(df) > 0 and str(df.iloc[0, 0]).startswith('E.g'):
            df = df.iloc[1:]  # Skip hint row
            df = df.reset_index(drop=True)
        
        # Normalize column names
        df.columns = df.columns.str.strip().str.lower().str.replace('*', '', regex=False)
        
        results = {"success": 0, "failed": 0, "errors": [], "medicines_deducted": {}}
        
        # Get all medicines for mapping
        medicines = await db.medicines.find({}, {"_id": 0}).to_list(None)
        medicine_map = {m["name"]: m for m in medicines}
        
        for idx, row in df.iterrows():
            row_num = idx + 3  # Account for header and hint rows
            try:
                # Validate required fields
                case_number = str(row.get('case number', '')).strip()
                if not case_number or case_number == 'nan':
                    results["errors"].append(f"Row {row_num}: Missing case number")
                    results["failed"] += 1
                    continue
                
                # Find the case
                case = await db.cases.find_one({"case_number": case_number}, {"_id": 0})
                if not case:
                    results["errors"].append(f"Row {row_num}: Case {case_number} not found")
                    results["failed"] += 1
                    continue
                
                # Check if surgery already exists
                if case.get("surgery"):
                    results["errors"].append(f"Row {row_num}: Case {case_number} already has surgery record")
                    results["failed"] += 1
                    continue
                
                # Parse date
                date_str = str(row.get('surgery date (dd/mm/yyyy)', '')).strip()
                try:
                    if isinstance(row.get('surgery date (dd/mm/yyyy)'), datetime):
                        surgery_date = row.get('surgery date (dd/mm/yyyy)')
                    else:
                        surgery_date = datetime.strptime(date_str, "%d/%m/%Y")
                except (ValueError, TypeError):
                    try:
                        surgery_date = datetime.strptime(date_str.split()[0], "%Y-%m-%d")
                    except (ValueError, TypeError):
                        results["errors"].append(f"Row {row_num}: Invalid date format. Use DD/MM/YYYY")
                        results["failed"] += 1
                        continue
                
                surgery_date = surgery_date.replace(tzinfo=timezone.utc)
                
                # Validate gender
                gender = str(row.get('gender', '')).strip().title()
                if gender not in ['Male', 'Female']:
                    results["errors"].append(f"Row {row_num}: Invalid gender. Use 'Male' or 'Female'")
                    results["failed"] += 1
                    continue
                
                # Validate weight
                try:
                    weight = float(row.get('weight (kg)', 0))
                    if weight < 10 or weight > 30:
                        results["errors"].append(f"Row {row_num}: Weight must be between 10-30 kg")
                        results["failed"] += 1
                        continue
                except (ValueError, TypeError):
                    results["errors"].append(f"Row {row_num}: Invalid weight format")
                    results["failed"] += 1
                    continue
                
                # Check if cancelled
                cancelled = str(row.get('surgery cancelled', 'No')).strip().lower()
                is_cancelled = cancelled in ['yes', 'y', 'true', '1']
                
                cancellation_reason = None
                if is_cancelled:
                    cancellation_reason = str(row.get('cancellation reason', '')).strip()
                    if not cancellation_reason or cancellation_reason == 'nan':
                        results["errors"].append(f"Row {row_num}: Cancellation reason is required when surgery is cancelled")
                        results["failed"] += 1
                        continue
                
                skin = str(row.get('skin condition', 'Normal')).strip()
                if skin == 'nan':
                    skin = 'Normal'
                
                remarks = str(row.get('remarks', '')).strip()
                if remarks == 'nan':
                    remarks = None
                
                # Calculate and deduct medicines if not cancelled
                medicines_used = {}
                medicines_to_deduct = {}
                
                if not is_cancelled:
                    for med_name, protocol in MEDICINE_PROTOCOL.items():
                        dosage = calculate_medicine_dosage(weight, med_name, gender)
                        if dosage > 0 and med_name in medicine_map:
                            med_id = medicine_map[med_name]["id"]
                            medicines_used[med_name] = dosage
                            medicines_to_deduct[med_id] = dosage
                            
                            # Deduct from stock
                            await db.medicines.update_one(
                                {"id": med_id},
                                {"$inc": {"current_stock": -dosage}}
                            )
                            
                            # Track deductions
                            if med_name not in results["medicines_deducted"]:
                                results["medicines_deducted"][med_name] = 0
                            results["medicines_deducted"][med_name] += dosage
                
                # Build surgery record
                surgery = {
                    "surgery_date": surgery_date.isoformat(),
                    "pre_surgery_status": "Cancel Surgery" if is_cancelled else "Fit for Surgery",
                    "cancellation_reason": cancellation_reason,
                    "surgery_type": "Ovariohysterectomy" if gender == "Female" else "Castration",
                    "weight": weight,
                    "skin": skin,
                    "anesthesia_used": [],
                    "complications": False,
                    "post_surgery_status": None if is_cancelled else "Good",
                    "veterinary_signature": f"{current_user.get('first_name', '')} {current_user.get('last_name', '')}",
                    "remarks": remarks,
                    "veterinary_id": current_user["id"],
                    "photo_links": [],
                    "medicines_used": medicines_used,
                    "bulk_uploaded": True
                }
                
                # Determine status
                new_status = CaseStatus.SURGERY_CANCELLED.value if is_cancelled else CaseStatus.SURGERY_COMPLETED.value
                
                # Update case with surgery record and initial observation if missing
                update_data = {
                    "surgery": surgery,
                    "status": new_status,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                
                # Add initial observation if not present (for bulk uploaded catching records)
                if not case.get("initial_observation"):
                    update_data["initial_observation"] = {
                        "gender": gender,
                        "kennel_number": None,
                        "approximate_age": "Adult 2-8 years",
                        "color_markings": "Unknown",
                        "body_condition": "Normal",
                        "temperament": "Calm",
                        "visible_injuries": False,
                        "catcher_id": current_user["id"],
                        "observation_date": surgery_date.isoformat(),
                        "bulk_uploaded": True
                    }
                
                await db.cases.update_one(
                    {"id": case["id"]},
                    {"$set": update_data}
                )
                
                results["success"] += 1
                
            except Exception as e:
                results["errors"].append(f"Row {row_num}: {str(e)}")
                results["failed"] += 1
        
        return {
            "message": f"Bulk upload completed. {results['success']} records processed, {results['failed']} failed.",
            "results": results
        }
        
    except Exception as e:
        logger.error(f"Bulk upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

# ==================== AUTO MEDICINE CALCULATION ENDPOINTS ====================

@api_router.get("/medicine-protocol")
async def get_medicine_protocol(current_user: dict = Depends(get_current_user)):
    """Get the default medicine protocol with dosage calculations"""
    return MEDICINE_PROTOCOL

@api_router.post("/calculate-medicines")
async def calculate_surgery_medicines(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Calculate medicine dosages for a surgery based on weight and gender"""
    weight = data.get("weight", 0)
    gender = data.get("gender", "Male")
    
    if weight < 10 or weight > 30:
        raise HTTPException(status_code=400, detail="Weight must be between 10-30 kg")
    
    # Get all medicines
    medicines = await db.medicines.find({}, {"_id": 0}).to_list(None)
    medicine_map = {m["name"]: m for m in medicines}
    
    calculated = {}
    for med_name in MEDICINE_PROTOCOL.keys():
        dosage = calculate_medicine_dosage(weight, med_name, gender)
        if med_name in medicine_map:
            med = medicine_map[med_name]
            calculated[med_name] = {
                "medicine_id": med["id"],
                "dosage": dosage,
                "unit": med["unit"],
                "current_stock": med["current_stock"],
                "sufficient_stock": med["current_stock"] >= dosage
            }
    
    return {
        "weight": weight,
        "gender": gender,
        "medicines": calculated
    }

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()