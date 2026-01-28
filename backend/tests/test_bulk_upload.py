"""
ABC Program Management System - Bulk Upload & Medicine Calculation Tests
Tests for bulk upload endpoints and auto medicine calculation functionality
"""
import pytest
import requests
import os
import openpyxl
from io import BytesIO

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Test credentials
TEST_EMAIL = "manoj@janicestrust.org"
TEST_PASSWORD = "Kashid@25067"


@pytest.fixture(scope="module")
def auth_token():
    """Get authentication token"""
    response = requests.post(f"{BASE_URL}/api/auth/login", json={
        "email": TEST_EMAIL,
        "password": TEST_PASSWORD
    })
    assert response.status_code == 200
    return response.json()["access_token"]


@pytest.fixture(scope="module")
def auth_headers(auth_token):
    """Get auth headers"""
    return {"Authorization": f"Bearer {auth_token}"}


class TestMedicineProtocol:
    """Medicine protocol endpoint tests"""
    
    def test_get_medicine_protocol(self, auth_headers):
        """Test getting medicine protocol"""
        response = requests.get(f"{BASE_URL}/api/medicine-protocol", headers=auth_headers)
        assert response.status_code == 200
        data = response.json()
        
        # Verify protocol structure
        assert "Anti-Rabies Vaccine" in data
        assert "Xylazine" in data
        assert "Ketamine" in data
        assert "Vicryl 2" in data  # Female-only medicine
        
        # Verify protocol fields
        assert data["Anti-Rabies Vaccine"]["fixed"] == True
        assert data["Xylazine"]["per_10kg"] == True
        assert data["Vicryl 2"]["female_only"] == True
        print(f"Medicine protocol has {len(data)} medicines")


class TestCalculateMedicines:
    """Medicine calculation endpoint tests"""
    
    def test_calculate_medicines_male(self, auth_headers):
        """Test medicine calculation for male animal"""
        response = requests.post(f"{BASE_URL}/api/calculate-medicines", 
            headers=auth_headers,
            json={"weight": 15, "gender": "Male"}
        )
        assert response.status_code == 200
        data = response.json()
        
        assert data["weight"] == 15
        assert data["gender"] == "Male"
        assert "medicines" in data
        
        # Verify dosage calculations
        medicines = data["medicines"]
        assert medicines["Anti-Rabies Vaccine"]["dosage"] == 1  # Fixed dose
        assert medicines["Xylazine"]["dosage"] == 1.5  # 1 * 15/10 = 1.5
        assert medicines["Vicryl 2"]["dosage"] == 0  # Female-only, should be 0 for male
        print(f"Calculated {len(medicines)} medicines for 15kg male")
    
    def test_calculate_medicines_female(self, auth_headers):
        """Test medicine calculation for female animal"""
        response = requests.post(f"{BASE_URL}/api/calculate-medicines", 
            headers=auth_headers,
            json={"weight": 20, "gender": "Female"}
        )
        assert response.status_code == 200
        data = response.json()
        
        assert data["weight"] == 20
        assert data["gender"] == "Female"
        
        medicines = data["medicines"]
        assert medicines["Xylazine"]["dosage"] == 2  # 1 * 20/10 = 2
        assert medicines["Vicryl 2"]["dosage"] == 0.2  # Female-only, should have value
        print(f"Calculated {len(medicines)} medicines for 20kg female")
    
    def test_calculate_medicines_weight_too_low(self, auth_headers):
        """Test medicine calculation with weight below minimum"""
        response = requests.post(f"{BASE_URL}/api/calculate-medicines", 
            headers=auth_headers,
            json={"weight": 5, "gender": "Male"}
        )
        assert response.status_code == 400
        assert "Weight must be between 10-30 kg" in response.json()["detail"]
        print("Weight validation (too low) working correctly")
    
    def test_calculate_medicines_weight_too_high(self, auth_headers):
        """Test medicine calculation with weight above maximum"""
        response = requests.post(f"{BASE_URL}/api/calculate-medicines", 
            headers=auth_headers,
            json={"weight": 35, "gender": "Male"}
        )
        assert response.status_code == 400
        assert "Weight must be between 10-30 kg" in response.json()["detail"]
        print("Weight validation (too high) working correctly")


class TestBulkUploadTemplates:
    """Bulk upload template download tests"""
    
    def test_download_catching_template(self, auth_headers):
        """Test downloading catching template"""
        response = requests.get(f"{BASE_URL}/api/bulk-upload/template/catching", 
            headers=auth_headers
        )
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers.get("content-type", "")
        assert len(response.content) > 0
        
        # Verify it's a valid Excel file
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        assert ws.title == "Catching Records"
        assert "Case Number" in ws.cell(row=1, column=1).value
        print("Catching template downloaded and validated")
    
    def test_download_surgery_template(self, auth_headers):
        """Test downloading surgery template"""
        response = requests.get(f"{BASE_URL}/api/bulk-upload/template/surgery", 
            headers=auth_headers
        )
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers.get("content-type", "")
        assert len(response.content) > 0
        
        # Verify it's a valid Excel file
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        assert ws.title == "Surgery Records"
        assert "Case Number" in ws.cell(row=1, column=1).value
        assert "Weight" in ws.cell(row=1, column=4).value
        print("Surgery template downloaded and validated")
    
    def test_download_invalid_template_type(self, auth_headers):
        """Test downloading invalid template type"""
        response = requests.get(f"{BASE_URL}/api/bulk-upload/template/invalid", 
            headers=auth_headers
        )
        assert response.status_code == 400
        assert "Invalid template type" in response.json()["detail"]
        print("Invalid template type rejected correctly")


class TestBulkUploadCatching:
    """Bulk upload catching records tests"""
    
    def create_catching_excel(self, case_numbers):
        """Helper to create catching Excel file"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catching Records"
        
        headers = ["Case Number*", "Date (DD/MM/YYYY)*", "Time (HH:MM)*", "Latitude*", "Longitude*", "Address*", "Ward Number", "Remarks"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        hints = ["E.g., JAPP-001-2024", "E.g., 25/12/2024", "E.g., 14:30", "E.g., 19.0760", "E.g., 72.8777", "Full address", "Optional", "Optional notes"]
        for col, hint in enumerate(hints, 1):
            ws.cell(row=2, column=col, value=hint)
        
        for idx, case_num in enumerate(case_numbers, 3):
            ws.cell(row=idx, column=1, value=case_num)
            ws.cell(row=idx, column=2, value="28/01/2026")
            ws.cell(row=idx, column=3, value="10:30")
            ws.cell(row=idx, column=4, value="19.0760")
            ws.cell(row=idx, column=5, value="72.8777")
            ws.cell(row=idx, column=6, value=f"Test Address {idx}, Mumbai")
            ws.cell(row=idx, column=7, value=f"Ward {idx}")
            ws.cell(row=idx, column=8, value=f"Test remark {idx}")
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def test_bulk_upload_catching_success(self, auth_headers):
        """Test successful bulk upload of catching records"""
        import time
        timestamp = int(time.time())
        case_numbers = [f"TEST-PYTEST-C{timestamp}-001", f"TEST-PYTEST-C{timestamp}-002"]
        excel_file = self.create_catching_excel(case_numbers)
        
        response = requests.post(f"{BASE_URL}/api/bulk-upload/catching",
            headers=auth_headers,
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        assert response.status_code == 200
        data = response.json()
        
        assert data["results"]["success"] == 2
        assert data["results"]["failed"] == 0
        print(f"Bulk upload catching: {data['results']['success']} success, {data['results']['failed']} failed")
    
    def test_bulk_upload_catching_duplicate(self, auth_headers):
        """Test bulk upload with duplicate case numbers"""
        # Try to upload same case numbers again - use existing test case
        case_numbers = ["TEST-BULK-001"]  # Already exists from previous tests
        excel_file = self.create_catching_excel(case_numbers)
        
        response = requests.post(f"{BASE_URL}/api/bulk-upload/catching",
            headers=auth_headers,
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        assert response.status_code == 200
        data = response.json()
        
        assert data["results"]["failed"] == 1
        assert "already exists" in data["results"]["errors"][0]
        print("Duplicate case number rejected correctly")


class TestBulkUploadSurgery:
    """Bulk upload surgery records tests"""
    
    def create_surgery_excel(self, records):
        """Helper to create surgery Excel file"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Surgery Records"
        
        headers = ["Case Number*", "Surgery Date (DD/MM/YYYY)*", "Gender*", "Weight (kg)*", "Surgery Cancelled*", "Cancellation Reason", "Skin Condition", "Remarks"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        hints = ["E.g., JAPP-001-2024", "E.g., 26/12/2024", "Male or Female", "E.g., 15", "Yes or No", "Required if cancelled", "Normal, Rough, or Visible Infection", "Optional notes"]
        for col, hint in enumerate(hints, 1):
            ws.cell(row=2, column=col, value=hint)
        
        for idx, record in enumerate(records, 3):
            for col, value in enumerate(record, 1):
                ws.cell(row=idx, column=col, value=value)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def test_bulk_upload_surgery_success(self, auth_headers):
        """Test successful bulk upload of surgery records"""
        import time
        timestamp = int(time.time())
        
        # First create catching records
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catching Records"
        headers = ["Case Number*", "Date (DD/MM/YYYY)*", "Time (HH:MM)*", "Latitude*", "Longitude*", "Address*", "Ward Number", "Remarks"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        hints = ["E.g.", "E.g.", "E.g.", "E.g.", "E.g.", "Full", "Opt", "Opt"]
        for col, hint in enumerate(hints, 1):
            ws.cell(row=2, column=col, value=hint)
        
        case1 = f"TEST-SURG-{timestamp}-001"
        case2 = f"TEST-SURG-{timestamp}-002"
        
        ws.cell(row=3, column=1, value=case1)
        ws.cell(row=3, column=2, value="28/01/2026")
        ws.cell(row=3, column=3, value="10:30")
        ws.cell(row=3, column=4, value="19.0760")
        ws.cell(row=3, column=5, value="72.8777")
        ws.cell(row=3, column=6, value="Test Address 1")
        
        ws.cell(row=4, column=1, value=case2)
        ws.cell(row=4, column=2, value="28/01/2026")
        ws.cell(row=4, column=3, value="10:30")
        ws.cell(row=4, column=4, value="19.0760")
        ws.cell(row=4, column=5, value="72.8777")
        ws.cell(row=4, column=6, value="Test Address 2")
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        requests.post(f"{BASE_URL}/api/bulk-upload/catching",
            headers=auth_headers,
            files={"file": ("test.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        # Now upload surgery records
        records = [
            [case1, "28/01/2026", "Male", "15", "No", "", "Normal", "Test surgery"],
            [case2, "28/01/2026", "Female", "20", "No", "", "Normal", "Test surgery female"],
        ]
        excel_file = self.create_surgery_excel(records)
        
        response = requests.post(f"{BASE_URL}/api/bulk-upload/surgery",
            headers=auth_headers,
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        assert response.status_code == 200
        data = response.json()
        
        assert data["results"]["success"] == 2
        assert "medicines_deducted" in data["results"]
        assert len(data["results"]["medicines_deducted"]) > 0
        print(f"Bulk upload surgery: {data['results']['success']} success, medicines deducted: {list(data['results']['medicines_deducted'].keys())[:5]}...")
    
    def test_bulk_upload_surgery_cancelled(self, auth_headers):
        """Test bulk upload of cancelled surgery (no medicine deduction)"""
        import time
        timestamp = int(time.time())
        case_num = f"TEST-CANCEL-{timestamp}"
        
        # First create a new catching record
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catching Records"
        headers = ["Case Number*", "Date (DD/MM/YYYY)*", "Time (HH:MM)*", "Latitude*", "Longitude*", "Address*", "Ward Number", "Remarks"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        hints = ["E.g.", "E.g.", "E.g.", "E.g.", "E.g.", "Full", "Opt", "Opt"]
        for col, hint in enumerate(hints, 1):
            ws.cell(row=2, column=col, value=hint)
        ws.cell(row=3, column=1, value=case_num)
        ws.cell(row=3, column=2, value="28/01/2026")
        ws.cell(row=3, column=3, value="10:30")
        ws.cell(row=3, column=4, value="19.0760")
        ws.cell(row=3, column=5, value="72.8777")
        ws.cell(row=3, column=6, value="Test Address Cancel")
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        requests.post(f"{BASE_URL}/api/bulk-upload/catching",
            headers=auth_headers,
            files={"file": ("test.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        # Now upload cancelled surgery
        records = [
            [case_num, "28/01/2026", "Male", "15", "Yes", "Too weak", "Rough", "Cancelled surgery"],
        ]
        excel_file = self.create_surgery_excel(records)
        
        response = requests.post(f"{BASE_URL}/api/bulk-upload/surgery",
            headers=auth_headers,
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        assert response.status_code == 200
        data = response.json()
        
        assert data["results"]["success"] == 1
        # Cancelled surgery should not deduct medicines
        assert len(data["results"]["medicines_deducted"]) == 0
        print("Cancelled surgery uploaded without medicine deduction")
    
    def test_bulk_upload_surgery_case_not_found(self, auth_headers):
        """Test bulk upload surgery for non-existent case"""
        records = [
            ["NON-EXISTENT-CASE", "28/01/2026", "Male", "15", "No", "", "Normal", "Test"],
        ]
        excel_file = self.create_surgery_excel(records)
        
        response = requests.post(f"{BASE_URL}/api/bulk-upload/surgery",
            headers=auth_headers,
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        assert response.status_code == 200
        data = response.json()
        
        assert data["results"]["failed"] == 1
        assert "not found" in data["results"]["errors"][0]
        print("Non-existent case rejected correctly")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
